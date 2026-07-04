import argparse
from datetime import time
from pathlib import Path
import re
import sys

BASE_DIR = Path(__file__).resolve().parents[1]
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from openpyxl import load_workbook

from db import DatabaseManager

IGNORAR_HOJAS = {"info", "datos2", "nuevo"}


def normalizar(texto):
    if texto is None:
        return ""
    texto = str(texto).strip().lower()
    reemplazos = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
        "ü": "u",
        "º": "",
        "°": "",
    }
    for origen, destino in reemplazos.items():
        texto = texto.replace(origen, destino)
    return " ".join(texto.split())


HEADER_ALIASES = {
    "pedido": ["no pedido", "n pedido", "nº pedido", "n° pedido"],
    "etiqueta": ["no de etiqueta", "n de etiqueta", "nºde etiqueta", "nº de etiqueta", "n°de etiqueta", "n° de etiqueta"],
    "caja": ["caja"],
    "lote": ["lote"],
    "producto": ["produccion", "producto"],
    "observaciones": ["observaciones"],
    "cantidad": ["cantidad", "cant."],
    "peso": ["peso kg", "peso"],
    "hora": ["hora fabr.", "hora fabr", "hora"],
}


def detectar_header(ws):
    for row in range(1, min(ws.max_row, 25) + 1):
        encontrados = {}
        for col in range(1, min(ws.max_column, 20) + 1):
            valor = normalizar(ws.cell(row, col).value)
            if not valor:
                continue

            if "pedido" in valor:
                encontrados["pedido"] = col
                continue
            if "etiqueta" in valor:
                encontrados["etiqueta"] = col
                continue
            if valor == "caja":
                encontrados["caja"] = col
                continue
            if valor == "lote":
                encontrados["lote"] = col
                continue
            if valor in ("produccion", "producto"):
                encontrados["producto"] = col
                continue
            if "observaciones" in valor:
                encontrados["observaciones"] = col
                continue
            if valor in ("cantidad", "cant."):
                encontrados["cantidad"] = col
                continue
            if "peso" in valor:
                encontrados["peso"] = col
                continue
            if "hora" in valor:
                encontrados["hora"] = col
                continue

            for clave, aliases in HEADER_ALIASES.items():
                if valor in aliases:
                    encontrados[clave] = col
        if {"pedido", "etiqueta", "cantidad"}.issubset(encontrados):
            return row, encontrados
    return None, {}


def parsear_nombre_fecha(nombre_hoja):
    match = re.match(r"^(\d{1,2})-(\d{1,2})(?:-(\d{2,4}))?$", nombre_hoja)
    if not match:
        return None
    dia = int(match.group(1))
    mes = int(match.group(2))
    year_raw = match.group(3)
    if not year_raw:
        return dia, mes, None
    year = int(year_raw)
    if len(year_raw) == 2:
        year += 2000
    return dia, mes, year


def construir_mapa_fechas(sheetnames):
    fechas = {}
    current_year = None
    prev_mes = None
    prev_dia = None

    for nombre in sheetnames:
        parsed = parsear_nombre_fecha(nombre)
        if not parsed:
            continue

        dia, mes, year = parsed
        if year is not None:
            current_year = year
        elif current_year is not None and prev_mes is not None and prev_mes >= 10 and mes <= 2:
            current_year += 1

        fechas[nombre] = f"{dia:02d}/{mes:02d}/{current_year}" if current_year is not None else f"{dia:02d}/{mes:02d}"
        prev_dia = dia
        prev_mes = mes

    return fechas


BOBINA_CLAVES = {
    "ancavico": "Bob. Ancavico",
    "transp": "Bob. Transparente",
    "transparente": "Bob. Transparente",
    "rejo": "Bob. Rejo",
    "frozen": "Bob. Frozzenbox",
    "martinez": "Bob. Martinez",
    "metal": "Plastico a Granel",
}


def mapear_bobina(observacion):
    texto = normalizar(observacion)
    for clave, valor in BOBINA_CLAVES.items():
        if clave in texto:
            return valor
    return str(observacion or "").strip()


def convertir_valor(valor):
    if valor is None:
        return ""
    if isinstance(valor, time):
        return valor.strftime("%H:%M")
    if isinstance(valor, float):
        entero = int(valor)
        return str(entero) if valor == entero else f"{valor:.2f}".rstrip("0").rstrip(".")
    return str(valor).strip()


def to_float(valor):
    try:
        return float(str(valor).replace(",", ".").strip() or 0)
    except ValueError:
        return 0.0


def extraer_incidencias_preheader(ws, header_row):
    incidencias = []
    for row in range(max(1, header_row - 6), header_row):
        etiqueta = convertir_valor(ws.cell(row, 1).value)
        if not etiqueta:
            continue
        texto = normalizar(etiqueta)
        if "bolsas" not in texto and "metal" not in texto:
            continue
        valores = []
        for col in range(2, min(ws.max_column, 8) + 1):
            celda = convertir_valor(ws.cell(row, col).value)
            if celda:
                valores.append(celda)
        if valores:
            incidencias.append(f"{etiqueta}: {' | '.join(valores)}")
    return " ; ".join(incidencias)


def extraer_lote_bobina(ws, header_row):
    pares = []
    for row in range(max(1, header_row - 6), header_row):
        etiqueta = convertir_valor(ws.cell(row, 1).value)
        if not etiqueta or normalizar(etiqueta) == "lote de bobina":
            continue
        valor = ""
        for col in range(2, min(ws.max_column, 8) + 1):
            celda = convertir_valor(ws.cell(row, col).value)
            if celda:
                valor = celda
                break
        pares.extend([etiqueta, valor])
        if len(pares) >= 6:
            break
    pares += [""] * (6 - len(pares))
    return pares[:6]


def extraer_lineas(ws, header_row, columnas):
    lineas = []
    for row in range(header_row + 1, ws.max_row + 1):
        pedido = convertir_valor(ws.cell(row, columnas.get("pedido", 0)).value) if columnas.get("pedido") else ""
        etiqueta = convertir_valor(ws.cell(row, columnas.get("etiqueta", 0)).value) if columnas.get("etiqueta") else ""
        caja = convertir_valor(ws.cell(row, columnas.get("caja", 0)).value) if columnas.get("caja") else ""
        lote = convertir_valor(ws.cell(row, columnas.get("lote", 0)).value) if columnas.get("lote") else ""
        producto = convertir_valor(ws.cell(row, columnas.get("producto", 0)).value) if columnas.get("producto") else ""
        observaciones = convertir_valor(ws.cell(row, columnas.get("observaciones", 0)).value) if columnas.get("observaciones") else ""
        cantidad = convertir_valor(ws.cell(row, columnas.get("cantidad", 0)).value) if columnas.get("cantidad") else ""
        peso = convertir_valor(ws.cell(row, columnas.get("peso", 0)).value) if columnas.get("peso") else ""
        hora = convertir_valor(ws.cell(row, columnas.get("hora", 0)).value) if columnas.get("hora") else ""

        if not any([pedido, etiqueta, caja, lote, producto, observaciones, cantidad, peso, hora]):
            continue

        es_bobina = not producto and "bobina" in normalizar(observaciones)
        tipo = "BOBINA" if es_bobina else "PALET"
        bobina = mapear_bobina(observaciones) if es_bobina else ""

        lineas.append(
            {
                "pedido": pedido,
                "tipo": tipo,
                "bobina": bobina,
                "etiqueta": etiqueta,
                "caja": caja,
                "lote": lote,
                "producto": producto,
                "observaciones": observaciones if observaciones else bobina,
                "cantidad": cantidad,
                "peso": to_float(peso),
                "palets": 1 if tipo == "PALET" and any([cantidad, peso, etiqueta, producto]) else 0,
                "fabricacion": observaciones if tipo == "PALET" else "",
                "ultima_hora": hora,
                "factor_producto": 0,
            }
        )
    return lineas


def construir_estado(ws, fecha_hoja):
    header_row, columnas = detectar_header(ws)
    if not header_row:
        return None

    lineas = extraer_lineas(ws, header_row, columnas)
    if not lineas:
        return None

    detalle_dia = {
        "fecha": fecha_hoja,
        "encargado": "",
        "operarios": "",
        "incidencias": extraer_incidencias_preheader(ws, header_row),
    }
    detalle_produccion = {
        "lote_bobina": extraer_lote_bobina(ws, header_row),
        "filas": [],
    }
    return {
        "lineas_produccion": lineas,
        "detalle_dia": detalle_dia,
        "detalle_produccion": detalle_produccion,
    }


def hojas_por_defecto(wb):
    seleccionadas = []
    for nombre in wb.sheetnames:
        if normalizar(nombre) in IGNORAR_HOJAS:
            continue
        if re.match(r"^\d{2}-\d{2}(-\d{2,4})?$", nombre):
            seleccionadas.append(nombre)
    return seleccionadas[:5]


def main():
    parser = argparse.ArgumentParser(description="Importador piloto Excel -> SQLite para producciones Ancavico")
    parser.add_argument("excel", help="Ruta al archivo .xlsm/.xlsx")
    parser.add_argument("--apply", action="store_true", help="Aplica la importacion en SQLite")
    parser.add_argument("--sheets", nargs="*", help="Nombres de hojas a importar en esta prueba")
    parser.add_argument("--force", action="store_true", help="Reimporta aunque la hoja ya estuviera registrada")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    wb = load_workbook(excel_path, data_only=True, keep_vba=True)
    nombres = args.sheets or hojas_por_defecto(wb)
    mapa_fechas = construir_mapa_fechas(wb.sheetnames)

    db = DatabaseManager()
    db.init_db()

    print(f"Archivo: {excel_path}")
    print(f"Modo: {'APLICAR' if args.apply else 'VISTA PREVIA'}")
    print(f"Hojas seleccionadas: {', '.join(nombres)}")
    print()

    importadas = 0
    omitidas = 0

    for nombre in nombres:
        if nombre not in wb.sheetnames:
            print(f"[OMITIDA] Hoja no encontrada: {nombre}")
            omitidas += 1
            continue

        estado = construir_estado(wb[nombre], mapa_fechas.get(nombre, nombre))
        if not estado:
            print(f"[OMITIDA] {nombre}: no se detecto un bloque importable")
            omitidas += 1
            continue

        source_file = str(excel_path.resolve())
        primera = estado['lineas_produccion'][0]
        if args.apply and not args.force:
            if db.has_imported_source(source_file, nombre):
                print(f"[OMITIDA] {nombre}: ya estaba importada anteriormente")
                omitidas += 1
                continue
            if db.has_production_signature(
                estado['detalle_dia']['fecha'],
                len(estado['lineas_produccion']),
                primera.get('pedido', ''),
                primera.get('etiqueta', ''),
                primera.get('producto', ''),
            ):
                print(f"[OMITIDA] {nombre}: coincide con una produccion ya guardada")
                omitidas += 1
                continue

        print(f"[OK] {nombre} -> fecha {estado['detalle_dia']['fecha']} | lineas: {len(estado['lineas_produccion'])}")
        print(
            f"     ejemplo: pedido={primera['pedido']} tipo={primera['tipo']} producto={primera['producto']} cantidad={primera['cantidad']} peso={primera['peso']}"
        )
        if estado['detalle_dia'].get('incidencias'):
            print(f"     incidencia detectada: {estado['detalle_dia']['incidencias']}")
        if args.apply:
            produccion_id = db.create_production(estado['detalle_dia']['fecha'])
            db.save_snapshot(produccion_id, estado)
            db.register_import_source(produccion_id, source_file, nombre)
            print(f"     guardada en SQLite con id {produccion_id}")
        importadas += 1

    print()
    print(f"Resumen -> importables: {importadas} | omitidas: {omitidas}")


if __name__ == "__main__":
    main()
